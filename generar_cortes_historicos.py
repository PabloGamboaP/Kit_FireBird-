"""
GENERADOR DE REPORTES DE CORTES HISTORICOS
===========================================

Genera reportes Excel de activos a fechas de corte específicas (31 dic de cada año)
Separa los activos en dos hojas: BODEGA y SERVICIOS

Uso:
    python generar_cortes_historicos.py

El script preguntará:
    1. Qué años desea generar (2022, 2023, ambos, u otro año)
    2. Dónde guardar los archivos (escritorio, carpeta actual, u otra ubicación)

Requisitos:
    - Archivo configuracion.ini con datos de conexión
    - Librería fdb instalada (pip install fdb)
    - Librería openpyxl instalada (pip install openpyxl)
"""

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    import fdb
except ImportError:
    print("ERROR: Librería 'fdb' no instalada")
    print("Instalar con: pip install fdb")
    exit(1)

try:
    import configparser
except ImportError:
    print("ERROR: Librería 'configparser' no instalada")
    exit(1)

# Códigos de servicio que van en la hoja BODEGA
CODIGOS_BODEGA = [
    '0100-ABNE',
    '0101-ABP',
    '0200-AGBNE',
    '0300-INS',
    '0501-TNE',
    '0502-TMN',
    '0600-BC',
    'BP001'
]

def cargar_configuracion():
    """Carga la configuración desde configuracion.ini"""
    config = configparser.ConfigParser()
    
    if not os.path.exists('configuracion.ini'):
        print("ERROR: No se encontró el archivo 'configuracion.ini'")
        print("Copia 'configuracion.ini.example' a 'configuracion.ini' y configúralo")
        return None
    
    config.read('configuracion.ini', encoding='utf-8')
    
    if 'FIREBIRD' not in config:
        print("ERROR: Sección [FIREBIRD] no encontrada en configuracion.ini")
        return None
    
    return config['FIREBIRD']

def get_connection():
    """Obtiene conexión a Firebird usando configuracion.ini"""
    try:
        fb_config = cargar_configuracion()
        if not fb_config:
            return None
        
        host = fb_config.get('HOST', '').strip()
        database = fb_config.get('DATABASE_PATH', '').strip()
        user = fb_config.get('USER', 'SYSDBA').strip()
        password = fb_config.get('PASSWORD', 'masterkey').strip()
        
        if not database:
            print("ERROR: DATABASE_PATH no configurado en configuracion.ini")
            return None
        
        # Usar fbclient.dll local
        FBCLIENT = os.path.join(os.path.dirname(__file__), "fbclient.dll")
        
        if not os.path.exists(FBCLIENT):
            print(f"ERROR: No se encontró fbclient.dll en {os.path.dirname(__file__)}")
            return None
        
        kwargs = dict(
            database=database,
            user=user,
            password=password,
            charset='WIN1252',
            fb_library_name=FBCLIENT,
        )
        
        if host:
            kwargs['host'] = host
        
        return fdb.connect(**kwargs)
        
    except Exception as e:
        print(f"ERROR conectando a Firebird: {str(e)}")
        return None

def consultar_activos_para_excel(ano):
    """Consulta activos al 31 dic del año especificado"""
    con = get_connection()
    if not con:
        return None
    
    try:
        cur = con.cursor()
        
        query = """
        SELECT 
            m.CODIGO AS PLACA,
            m.DESMAT AS DESCRIPCION,
            m.VALORCOMP AS VALOR_COMPRA,
            gc.CODIGO AS COD_GRUPO_CONTABLE,
            gc.DESCRIP AS NOM_GRUPO_CONTABLE,
            pc.CODIGO AS CUENTA_PUC,
            se.CODSERVI AS COD_SERVICIO,
            se.NOMSERVI AS NOM_SERVICIO,
            t.NIT AS CEDULA,
            t.NOMBRE AS RESPONSABLE
        FROM MATERIAL m 
        INNER JOIN SALAJUSTES sa ON m.MATID = sa.MATID
        LEFT JOIN GCMAT gc ON sa.GCMATID = gc.GCMATID
        LEFT JOIN PLANCUENTAS pc ON gc.CTA_INV = pc.PUCID
        LEFT JOIN SERVICIO se ON sa.SERVICIOID = se.SERVICIOID
        LEFT JOIN TERCEROS t ON sa.RESPONSABLE = t.TERID
        WHERE 
            sa.ANO = ?
            AND sa.MES = '12'
            AND (m.FEC_BAJA IS NULL OR m.FEC_BAJA > ?)
        ORDER BY 
            se.CODSERVI,
            t.NOMBRE,
            m.CODIGO
        """
        
        fecha_corte = datetime(ano, 12, 31, 23, 59, 59)
        cur.execute(query, [ano, fecha_corte])
        
        activos = []
        for row in cur.fetchall():
            activos.append({
                'placa': row[0],
                'descripcion': row[1],
                'valor_compra': row[2] or 0,
                'cod_grupo_contable': row[3],
                'nom_grupo_contable': row[4],
                'cuenta_puc': row[5],
                'cod_servicio': row[6],
                'nom_servicio': row[7],
                'cedula': row[8],
                'responsable': row[9]
            })
        
        cur.close()
        con.close()
        
        return activos
        
    except Exception as e:
        print(f"ERROR consultando activos: {str(e)}")
        import traceback
        traceback.print_exc()
        if con:
            con.close()
        return None

def separar_por_bodega(activos):
    """Separa activos en BODEGA y SERVICIOS"""
    bodega = []
    servicios = []
    
    for activo in activos:
        cod_servicio = activo['cod_servicio'] or ''
        
        if cod_servicio in CODIGOS_BODEGA:
            bodega.append(activo)
        else:
            servicios.append(activo)
    
    return bodega, servicios

def crear_excel(bodega, servicios, nombre_archivo, ano):
    """Crea archivo Excel con 2 hojas: BODEGA y SERVICIOS"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    nota_font = Font(bold=True, size=12, color="000000")
    nota_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    nota_alignment = Alignment(horizontal="left", vertical="center")
    
    # Columnas
    columnas = [
        ('PLACA', 15),
        ('DESCRIPCION', 50),
        ('VALOR COMPRA', 15),
        ('COD GRUPO CONTABLE', 20),
        ('GRUPO CONTABLE', 30),
        ('CUENTA PUC', 15),
        ('COD SERVICIO', 15),
        ('SERVICIO', 40),
        ('CEDULA', 15),
        ('RESPONSABLE', 40)
    ]
    
    total_bienes = len(bodega) + len(servicios)
    
    # Crear hoja BODEGA
    ws_bodega = wb.create_sheet("BODEGA")
    
    nota_bodega = f"REPORTE DE ACTIVOS - CORTE AL 31 DE DICIEMBRE DE {ano} | Total de bienes: {total_bienes:,} | Bienes en BODEGA: {len(bodega):,}"
    ws_bodega.merge_cells('A1:J1')
    cell_nota = ws_bodega['A1']
    cell_nota.value = nota_bodega
    cell_nota.font = nota_font
    cell_nota.fill = nota_fill
    cell_nota.alignment = nota_alignment
    ws_bodega.row_dimensions[1].height = 25
    
    for col_num, (nombre, ancho) in enumerate(columnas, 1):
        cell = ws_bodega.cell(row=2, column=col_num)
        cell.value = nombre
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws_bodega.column_dimensions[get_column_letter(col_num)].width = ancho
    
    for row_num, activo in enumerate(bodega, 3):
        ws_bodega.cell(row=row_num, column=1, value=activo['placa'])
        ws_bodega.cell(row=row_num, column=2, value=activo['descripcion'])
        ws_bodega.cell(row=row_num, column=3, value=activo['valor_compra'])
        ws_bodega.cell(row=row_num, column=4, value=activo['cod_grupo_contable'])
        ws_bodega.cell(row=row_num, column=5, value=activo['nom_grupo_contable'])
        ws_bodega.cell(row=row_num, column=6, value=activo['cuenta_puc'])
        ws_bodega.cell(row=row_num, column=7, value=activo['cod_servicio'])
        ws_bodega.cell(row=row_num, column=8, value=activo['nom_servicio'])
        ws_bodega.cell(row=row_num, column=9, value=activo['cedula'])
        ws_bodega.cell(row=row_num, column=10, value=activo['responsable'])
        ws_bodega.cell(row=row_num, column=3).number_format = '$#,##0.00'
    
    # Crear hoja SERVICIOS
    ws_servicios = wb.create_sheet("SERVICIOS")
    
    nota_servicios = f"REPORTE DE ACTIVOS - CORTE AL 31 DE DICIEMBRE DE {ano} | Total de bienes: {total_bienes:,} | Bienes en SERVICIOS: {len(servicios):,}"
    ws_servicios.merge_cells('A1:J1')
    cell_nota = ws_servicios['A1']
    cell_nota.value = nota_servicios
    cell_nota.font = nota_font
    cell_nota.fill = nota_fill
    cell_nota.alignment = nota_alignment
    ws_servicios.row_dimensions[1].height = 25
    
    for col_num, (nombre, ancho) in enumerate(columnas, 1):
        cell = ws_servicios.cell(row=2, column=col_num)
        cell.value = nombre
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws_servicios.column_dimensions[get_column_letter(col_num)].width = ancho
    
    for row_num, activo in enumerate(servicios, 3):
        ws_servicios.cell(row=row_num, column=1, value=activo['placa'])
        ws_servicios.cell(row=row_num, column=2, value=activo['descripcion'])
        ws_servicios.cell(row=row_num, column=3, value=activo['valor_compra'])
        ws_servicios.cell(row=row_num, column=4, value=activo['cod_grupo_contable'])
        ws_servicios.cell(row=row_num, column=5, value=activo['nom_grupo_contable'])
        ws_servicios.cell(row=row_num, column=6, value=activo['cuenta_puc'])
        ws_servicios.cell(row=row_num, column=7, value=activo['cod_servicio'])
        ws_servicios.cell(row=row_num, column=8, value=activo['nom_servicio'])
        ws_servicios.cell(row=row_num, column=9, value=activo['cedula'])
        ws_servicios.cell(row=row_num, column=10, value=activo['responsable'])
        ws_servicios.cell(row=row_num, column=3).number_format = '$#,##0.00'
    
    wb.save(nombre_archivo)
    print(f"\nArchivo Excel creado: {nombre_archivo}")
    print(f"  - Hoja BODEGA: {len(bodega):,} registros")
    print(f"  - Hoja SERVICIOS: {len(servicios):,} registros")
    print(f"  - Total: {len(bodega) + len(servicios):,} registros")

def generar_corte(ano, ruta_destino):
    """Genera el corte para un año específico"""
    print("=" * 80)
    print(f"REPORTE DE ACTIVOS AL 31 DICIEMBRE {ano}")
    print("=" * 80)
    print()
    
    print(f"Consultando activos al 31 dic {ano}...")
    activos = consultar_activos_para_excel(ano)
    
    if not activos:
        print(f"ERROR: No se pudieron consultar los activos de {ano}")
        return False
    
    print(f"Total de activos consultados: {len(activos):,}")
    print()
    
    print("Separando activos por BODEGA y SERVICIOS...")
    bodega, servicios = separar_por_bodega(activos)
    
    print(f"  - BODEGA: {len(bodega):,} activos")
    print(f"  - SERVICIOS: {len(servicios):,} activos")
    print()
    
    nombre_archivo = os.path.join(ruta_destino, f"Corte_{ano}.xlsx")
    
    print("Generando archivo Excel...")
    crear_excel(bodega, servicios, nombre_archivo, ano)
    
    print()
    return True

def main():
    print("=" * 80)
    print("GENERADOR DE REPORTES DE CORTES HISTORICOS")
    print("=" * 80)
    print()
    
    # Seleccionar años
    print("Selecciona qué años deseas generar:")
    print("  1. Solo 2022")
    print("  2. Solo 2023")
    print("  3. Ambos (2022 y 2023)")
    print("  4. Otro año (ingresar manualmente)")
    print()
    
    opcion = input("Opción (1-4): ").strip()
    
    anos = []
    if opcion == '1':
        anos = [2022]
    elif opcion == '2':
        anos = [2023]
    elif opcion == '3':
        anos = [2022, 2023]
    elif opcion == '4':
        ano_custom = input("Ingresa el año (ej: 2021): ").strip()
        try:
            anos = [int(ano_custom)]
        except:
            print("ERROR: Año inválido")
            return
    else:
        print("Opción inválida")
        return
    
    print()
    
    # Seleccionar ubicación
    print("=" * 80)
    print("UBICACION DE LOS ARCHIVOS")
    print("=" * 80)
    print()
    print("Opciones:")
    print("  1. Carpeta actual (Kit_FireBird)")
    print("  2. Escritorio")
    print("  3. Otra ubicación (ingresar ruta)")
    print()
    
    opcion_ruta = input("Selecciona una opción (1-3): ").strip()
    
    if opcion_ruta == '1':
        ruta_destino = os.path.dirname(__file__)
    elif opcion_ruta == '2':
        ruta_destino = os.path.join(os.path.expanduser('~'), 'Desktop')
    elif opcion_ruta == '3':
        ruta_destino = input("Ingresa la ruta completa: ").strip()
        if not os.path.exists(ruta_destino):
            print(f"\nERROR: La ruta '{ruta_destino}' no existe")
            return
    else:
        print("\nOpción inválida. Usando carpeta actual.")
        ruta_destino = os.path.dirname(__file__)
    
    print()
    
    # Generar cortes
    exitos = []
    for ano in anos:
        exito = generar_corte(ano, ruta_destino)
        exitos.append((ano, exito))
        if len(anos) > 1 and ano != anos[-1]:
            print()
    
    print()
    print("=" * 80)
    print("PROCESO COMPLETADO")
    print("=" * 80)
    print()
    
    for ano, exito in exitos:
        if exito:
            archivo = os.path.join(ruta_destino, f"Corte_{ano}.xlsx")
            print(f"Corte {ano}: {archivo}")
    
    print()

if __name__ == '__main__':
    main()
